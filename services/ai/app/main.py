from fastapi import FastAPI, HTTPException
from pydantic import BaseModel, Field
import re, math, base64, io, zipfile
from collections import Counter
from xml.etree import ElementTree as ET
from typing import Any
import csv
from html.parser import HTMLParser

app = FastAPI(title="ARIS AI Service", version="4.0.0")
class ChunkRequest(BaseModel):
    text:str=""; target_tokens:int=420; max_tokens:int=720; min_tokens:int=90; overlap_tokens:int=60
    preserve_headings:bool=True; preserve_lists:bool=True; preserve_tables:bool=True; event_aware:bool=True; topic_aware:bool=True
    blocks:list[dict[str,Any]]=[]; document_context:dict[str,Any]={}
class Heading(BaseModel): id:str; title:str; description:str=""
class AffinityRequest(BaseModel): text:str; headings:list[Heading]
class ParseRequest(BaseModel): file_name:str; content_base64:str=""; text:str=""; document_id:str|None=None

def tokens(text): return re.findall(r"[A-Za-z0-9']+", text.lower())
def cosine(a,b):
    ca,cb=Counter(tokens(a)),Counter(tokens(b)); common=set(ca)&set(cb); dot=sum(ca[x]*cb[x] for x in common); na=math.sqrt(sum(v*v for v in ca.values())); nb=math.sqrt(sum(v*v for v in cb.values())); return dot/(na*nb) if na and nb else 0.0

def classify_line(line):
    s=line.strip()
    if not s:return None
    if re.match(r'^(\d+(?:\.\d+)*)\s+\S',s): return ('heading', min(5,s.split()[0].count('.')+1))
    if len(s)<90 and (s.isupper() or s.istitle()): return ('heading',2)
    if re.match(r'^[-•*]\s+',s):return ('list',None)
    return ('paragraph',None)

def normalise_text(text):
    blocks=[]; order=0
    for raw in re.split(r'\n+',text):
        line=raw.strip(); c=classify_line(line)
        if not c:continue
        order+=1; typ,level=c
        if typ=='list': blocks.append({'id':f'blk-{order}','type':'list','items':[re.sub(r'^[-•*]\s+','',line)],'page':1,'order':order,'confidence':0.9})
        else: blocks.append({'id':f'blk-{order}','type':typ,'level':level,'text':line,'page':1,'order':order,'confidence':0.9 if typ=='heading' else 0.95})
    return blocks

def parse_docx(data):
    with zipfile.ZipFile(io.BytesIO(data)) as z: xml=z.read('word/document.xml')
    root=ET.fromstring(xml); ns={'w':'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}; blocks=[]; order=0
    for p in root.findall('.//w:p',ns):
        text=''.join(t.text or '' for t in p.findall('.//w:t',ns)).strip()
        if not text:continue
        order+=1; style=p.find('.//w:pStyle',ns); stylev=style.attrib.get('{%s}val'%ns['w'],'') if style is not None else ''
        m=re.search(r'Heading(\d+)',stylev,re.I); typ='heading' if m else 'paragraph'; b={'id':f'blk-{order}','type':typ,'text':text,'page':1,'order':order,'confidence':0.98}
        if m:b['level']=int(m.group(1))
        blocks.append(b)
    return blocks

def parse_pptx(data):
    out=[]; order=0
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        names=sorted(n for n in z.namelist() if n.startswith('ppt/slides/slide') and n.endswith('.xml'))
        for page,n in enumerate(names,1):
            root=ET.fromstring(z.read(n)); texts=[x.text or '' for x in root.iter() if x.tag.endswith('}t')]
            for i,t in enumerate(texts):
                if not t.strip():continue
                order+=1;out.append({'id':f'blk-{order}','type':'heading' if i==0 else 'paragraph','level':2 if i==0 else None,'text':t.strip(),'page':page,'order':order,'confidence':0.91})
    return out


def parse_pdf(data):
    try:
        from pypdf import PdfReader
    except ImportError: raise HTTPException(500,'pypdf is not installed')
    out=[];order=0
    for page_no,page in enumerate(PdfReader(io.BytesIO(data)).pages,1):
        for raw in (page.extract_text() or '').splitlines():
            c=classify_line(raw)
            if not c:continue
            order+=1;typ,level=c;out.append({'id':f'blk-{order}','type':typ,'level':level,'text':raw.strip(),'page':page_no,'order':order,'confidence':0.84})
    return out

def parse_xlsx(data):
    try:
        from openpyxl import load_workbook
    except ImportError: raise HTTPException(500,'openpyxl is not installed')
    wb=load_workbook(io.BytesIO(data),data_only=True);out=[];order=0
    for ws in wb.worksheets:
        order+=1;out.append({'id':f'blk-{order}','type':'heading','level':1,'text':ws.title,'page':1,'order':order,'confidence':1})
        rows=[[str(c) if c is not None else '' for c in r] for r in ws.iter_rows(values_only=True)]
        rows=[r for r in rows if any(x for x in r)]
        if rows:
            order+=1;out.append({'id':f'blk-{order}','type':'table','caption':ws.title,'columns':rows[0],'rows':rows[1:],'page':1,'order':order,'confidence':0.98})
    return out

def parse_rtf(data):
    text=data.decode('latin-1',errors='replace');text=re.sub(r'\\[a-z]+-?\d* ?',' ',text);text=re.sub(r'[{}]','',text);return normalise_text(text)

def extract_entities(text):
    return {'dates':list(dict.fromkeys(re.findall(r'\b(?:\d{1,2}\s+)?(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+20\d{2}\b',text)))[:20], 'metrics':list(dict.fromkeys(re.findall(r'\b\d+(?:\.\d+)?%|\b\d+[ ,]?\d*\s+(?:students|participants|sessions|events|publications)\b',text,re.I)))[:20]}

def block_text(b):
    if b.get('type')=='table':
        rows=[' | '.join(map(str,r)) for r in b.get('rows',[])]
        return ' '.join([b.get('caption',''),' | '.join(b.get('columns',[])),*rows]).strip()
    return (b.get('text') or ' '.join(b.get('items',[]))).strip()

def infer_type(text, block_types):
    if 'table' in block_types:return 'table'
    if 'list' in block_types:return 'list'
    if re.search(r'\b(?:organised|conducted|held|workshop|conference|event)\b',text,re.I) and re.search(r'\b20\d{2}\b|\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\b',text,re.I):return 'event'
    if re.search(r'\b\d+(?:\.\d+)?%|\b\d+\s+(?:students|participants|sessions|events)\b',text,re.I):return 'metric'
    return 'topic'

def structure_aware_chunks(req:ChunkRequest):
    blocks=req.blocks or normalise_text(req.text)
    path=[]; groups=[]; current=[]; current_tokens=0; current_path=[]
    def flush():
        nonlocal current,current_tokens,current_path
        if not current:return
        if not any(b.get('type')!='heading' for b in current):
            current=[]; current_tokens=0; return
        text='\n'.join(block_text(b) for b in current if block_text(b)).strip()
        if text: groups.append((current[:],current_path[:],text))
        current=[];current_tokens=0
    for b in sorted(blocks,key=lambda x:x.get('order',0)):
        if b.get('type')=='heading':
            flush(); lvl=max(1,int(b.get('level') or 1)); path=path[:lvl-1]+[block_text(b)]; current_path=path[:]
            if req.preserve_headings: current=[b];current_tokens=len(tokens(block_text(b)))
            continue
        txt=block_text(b); n=len(tokens(txt))
        atomic=(req.preserve_tables and b.get('type')=='table') or (req.preserve_lists and b.get('type')=='list')
        if current and current_tokens+n>req.target_tokens and (current_tokens>=req.min_tokens or atomic): flush(); current_path=path[:]
        if n>req.max_tokens and not atomic:
            for sentence in re.split(r'(?<=[.!?])\s+',txt):
                sn=len(tokens(sentence))
                if current and current_tokens+sn>req.target_tokens: flush();current_path=path[:]
                current.append({**b,'text':sentence});current_tokens+=sn
        else: current.append(b);current_tokens+=n
    flush()
    chunks=[]
    for idx,(bs,hpath,text) in enumerate(groups,1):
        ents=extract_entities(text); block_types=[b.get('type','paragraph') for b in bs]
        chunks.append({'id':f'generated-{idx}','text':text,'tokens':len(tokens(text)),'type':infer_type(text,block_types),'headingPath':hpath,'sourceSection':' / '.join(hpath),'blockIds':[b.get('id') for b in bs if b.get('id')],'topics':list(dict.fromkeys(re.findall(r'\b(?:student wellbeing|counselling|research|publications|governance|quality assurance|residential support|student engagement)\b',text,re.I))), 'dates':ents['dates'],'metrics':ents['metrics'],'context':req.document_context,'quality':94 if hpath else 86})
    return chunks

@app.get('/health')
def health():return {'status':'ok','version':'4.0.0','capabilities':['docx','pdf','xlsx','pptx','rtf','txt','csv','html','structural-normalisation','hierarchy-aware-chunking','event-aware-chunking','table-and-list-preservation']}
@app.post('/parse')
def parse(req:ParseRequest):
    data=base64.b64decode(req.content_base64) if req.content_base64 else b''; ext=req.file_name.lower().rsplit('.',1)[-1]
    try:
        if req.text: blocks=normalise_text(req.text)
        elif ext=='docx':blocks=parse_docx(data)
        elif ext=='pptx':blocks=parse_pptx(data)
        elif ext=='pdf':blocks=parse_pdf(data)
        elif ext in ('xlsx','xlsm'):blocks=parse_xlsx(data)
        elif ext=='rtf':blocks=parse_rtf(data)
        elif ext in ('txt','md','html','htm','csv'):blocks=normalise_text(data.decode('utf-8',errors='replace'))
        else: raise HTTPException(415,f'Parser for .{ext} requires an optional extraction adapter')
    except zipfile.BadZipFile: raise HTTPException(400,'The uploaded Office file is not a valid ZIP package')
    all_text=' '.join(block_text(b) for b in blocks)
    return {'documentId':req.document_id,'fileName':req.file_name,'blocks':blocks,'blockCount':len(blocks),'wordCount':len(tokens(all_text)),'entities':extract_entities(all_text),'warnings':[],'quality':94 if blocks else 0}
@app.post('/chunk')
def chunk(req:ChunkRequest): return {'chunks':structure_aware_chunks(req),'profile':{'targetTokens':req.target_tokens,'maxTokens':req.max_tokens,'minTokens':req.min_tokens,'overlapTokens':req.overlap_tokens}}
@app.post('/affinity')
def affinity(req:AffinityRequest):
    scores=[]
    for h in req.headings:
        score=min(.99,.15+cosine(req.text,h.title+' '+h.description)*1.6); assignment='Primary' if score>=.75 else 'Secondary' if score>=.55 else 'Review' if score>=.35 else 'None';scores.append({'headingId':h.id,'headingTitle':h.title,'score':round(score,2),'assignment':assignment})
    return {'mappings':sorted(scores,key=lambda x:x['score'],reverse=True)}
